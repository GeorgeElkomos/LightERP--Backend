import os
import django
from openpyxl import Workbook

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'erp_project.settings')
django.setup()

from Finance.budget_control.models import BudgetHeader, BudgetSegmentValue
from Finance.GL.models import XX_Segment

# Get budget ID 2
budget = BudgetHeader.objects.get(id=2)
print(f"Budget: {budget.budget_code} - {budget.budget_name}")
print(f"Status: {budget.status}")

# Get segments associated with this budget
segment_values = BudgetSegmentValue.objects.filter(budget=budget).select_related('segment_value')
print(f"\nSegments in budget:")
for bsv in segment_values:
    seg = bsv.segment_value
    print(f"  - {seg.code}: {seg.alias} (Type: {seg.segment_type.segment_name})")

# Create sample Excel file with actual segment codes
wb = Workbook()
ws = wb.active
ws.title = 'Budget Import Template'

# Headers
headers = ['Segment Code', 'Original Budget', 'Adjustment', 'Notes']
for col_idx, header in enumerate(headers, start=1):
    ws.cell(row=1, column=col_idx, value=header)

# Add actual segment data
row = 2
for bsv in segment_values:
    seg = bsv.segment_value
    ws.cell(row=row, column=1, value=seg.code)
    ws.cell(row=row, column=2, value=100000.00)  # Original Budget
    ws.cell(row=row, column=3, value=0.00)       # Adjustment
    ws.cell(row=row, column=4, value=f'Budget for {seg.alias}')
    row += 1

# Save file
filename = 'budget_import_sample.xlsx'
wb.save(filename)
print(f"\n✓ Created sample file: {filename}")
print(f"  Contains {row-2} segment(s)")
print(f"\nYou can now:")
print(f"1. Edit the 'Original Budget' amounts in the Excel file")
print(f"2. Import using: POST {{{{base_url}}}}/finance/budget/budget-headers/2/import/")
print(f"   - Method: POST")
print(f"   - Body: form-data")
print(f"   - Key: file (type: File)")
print(f"   - Value: Select the Excel file")
