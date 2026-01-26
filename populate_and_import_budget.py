import os
import django
from openpyxl import load_workbook

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'erp_project.settings')
django.setup()

from Finance.budget_control.models import BudgetHeader, BudgetSegmentValue
from Finance.budget_control.excel_utils import import_budget_from_excel

# Get budget ID 2
budget = BudgetHeader.objects.get(id=2)
print(f"Budget: {budget.budget_code} - {budget.budget_name}")
print(f"Status: {budget.status}")

# Get segments associated with this budget
segment_values = BudgetSegmentValue.objects.filter(budget_header=budget).select_related('segment_value')
print(f"\nSegments in budget: {segment_values.count()}")

if segment_values.count() == 0:
    print("⚠️  No segments found in this budget. Need to add segments first!")
    print("\nTo add segments, use the API:")
    print("POST /finance/budget/budget-headers/2/segments/")
    print('Body: {"segment_value_id": <segment_id>}')
else:
    # Load the template file
    template_file = 'Budget_Import_Template_BUD-2026-Q2.xlsx'
    
    try:
        wb = load_workbook(template_file)
        ws = wb.active
        
        print(f"\n✓ Loaded template: {template_file}")
        
        # Clear existing data (keep headers in row 1)
        ws.delete_rows(2, ws.max_row)
        
        # Add actual segment data
        row = 2
        for bsv in segment_values:
            seg = bsv.segment_value
            print(f"  Adding: {seg.code} - {seg.alias}")
            ws.cell(row=row, column=1, value=seg.code)
            ws.cell(row=row, column=2, value=100000.00)  # Original Budget
            ws.cell(row=row, column=3, value=0.00)       # Adjustment
            ws.cell(row=row, column=4, value=f'Budget for {seg.alias}')
            row += 1
        
        # Save updated file
        updated_file = 'Budget_Import_Ready_BUD-2026-Q2.xlsx'
        wb.save(updated_file)
        print(f"\n✓ Created populated file: {updated_file}")
        print(f"  Contains {row-2} segment(s) with 100,000.00 each")
        
        # Now import it
        print(f"\n📤 Importing budget amounts...")
        with open(updated_file, 'rb') as f:
            results = import_budget_from_excel(budget, f)
        
        if 'error' in results:
            print(f"✗ Import failed: {results['error']}")
        else:
            print(f"✓ Import successful!")
            print(f"  Total rows: {results['total_rows']}")
            print(f"  Success: {results['success_count']}")
            print(f"  Errors: {results['error_count']}")
            if results['error_count'] > 0:
                print(f"\n  Error details:")
                for error in results.get('errors', []):
                    print(f"    - {error}")
            
            # Show budget amounts
            print(f"\n📊 Current budget amounts:")
            from Finance.budget_control.models import BudgetAmount
            amounts = BudgetAmount.objects.filter(
                budget_segment_value__budget_header=budget
            ).select_related('budget_segment_value__segment_value')
            
            for amount in amounts:
                seg = amount.budget_segment_value.segment_value
                control = amount.get_effective_control_level()
                print(f"  {seg.code}: {amount.original_budget:,.2f} (Control: {control})")
                
    except FileNotFoundError:
        print(f"\n✗ Template file not found: {template_file}")
        print("Please ensure the file is in the current directory")
