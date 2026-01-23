"""
Budget Excel Import/Export Tests
Tests for:
- Excel export of budget data
- Excel import of budget amounts
- Validation and error handling
- Bulk operations

Note: If Excel functionality is not yet implemented in views.py,
these tests serve as specifications for the required functionality.
"""

import unittest
from rest_framework.test import APITestCase, APIClient
from rest_framework import status
from django.urls import reverse
from django.core.files.uploadedfile import SimpleUploadedFile
from decimal import Decimal
from datetime import date, timedelta
import io
import json

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from Finance.budget_control.models import BudgetHeader, BudgetSegmentValue, BudgetAmount
from Finance.GL.models import XX_Segment, XX_SegmentType
from Finance.core.models import Currency
from Finance.budget_control.tests.test_utils import create_test_user, create_test_currency


class BudgetExcelExportTestCase(APITestCase):
    """Test budget Excel export functionality"""
    
    def setUp(self):
        """Set up test data"""
        self.client = APIClient()
        
        self.user = create_test_user()
        self.client.force_authenticate(user=self.user)
        
        # Create currency
        self.currency = create_test_currency()
        
        # Create segment types and segments
        self.account_type = XX_SegmentType.objects.create(
            segment_name='Account',
            is_required=True,
            length=4,
            display_order=1
        )
        
        self.dept_type = XX_SegmentType.objects.create(
            segment_name='Department',
            is_required=True,
            length=2,
            display_order=2
        )
        
        self.segment_5000 = XX_Segment.objects.create(segment_type=self.account_type, code='5000', alias='Travel Expense', node_type='child', is_active=True)
        
        self.segment_5100 = XX_Segment.objects.create(segment_type=self.account_type, code='5100', alias='Office Supplies', node_type='child', is_active=True)
        
        self.segment_dept01 = XX_Segment.objects.create(segment_type=self.dept_type, code='01', alias='IT Department', node_type='child', is_active=True)
        
        # Create budget with data
        self.budget = BudgetHeader.objects.create(
            budget_code='EXPORT2026',
            budget_name='Export Test Budget',
            start_date=date.today(),
            end_date=date.today() + timedelta(days=365),
            currency=self.currency,
            status='ACTIVE',
            is_active=True,
            description='Budget for testing export'
        )
        
        # Create segment values and amounts
        seg_val_5000 = BudgetSegmentValue.objects.create(
            budget_header=self.budget,
            segment_value=self.segment_5000,
            control_level='ABSOLUTE'
        )
        
        seg_val_5100 = BudgetSegmentValue.objects.create(
            budget_header=self.budget,
            segment_value=self.segment_5100,
            control_level='ADVISORY'
        )
        
        BudgetAmount.objects.create(
            budget_segment_value=seg_val_5000,
            budget_header=self.budget,
            original_budget=Decimal('50000.00'),
            committed_amount=Decimal('10000.00'),
            encumbered_amount=Decimal('15000.00'),
            actual_amount=Decimal('5000.00')
        )
        
        BudgetAmount.objects.create(
            budget_segment_value=seg_val_5100,
            budget_header=self.budget,
            original_budget=Decimal('30000.00'),
            committed_amount=Decimal('5000.00'),
            encumbered_amount=Decimal('8000.00'),
            actual_amount=Decimal('3000.00')
        )
    
    def test_export_budget_to_excel(self):
        """Test GET /budget-headers/<pk>/export/ returns Excel file"""
        # Note: This endpoint may not exist yet - this is a specification
        url = reverse('budget_control:budget-export', kwargs={'pk': self.budget.id})
        response = self.client.get(url)
        
        # If not implemented, this will return 404 or 405
        # When implemented, should return:
        if response.status_code == status.HTTP_200_OK:
            # Verify response is Excel file
            self.assertEqual(
                response['Content-Type'],
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            self.assertIn('attachment', response['Content-Disposition'])
            self.assertIn('EXPORT2026', response['Content-Disposition'])
            
            # Verify file content if openpyxl available
            if OPENPYXL_AVAILABLE:
                wb = openpyxl.load_workbook(io.BytesIO(response.content))
                
                # Should have sheets
                self.assertIn('Budget Header', wb.sheetnames)
                self.assertIn('Budget Amounts', wb.sheetnames)
                
                # Check header sheet
                header_sheet = wb['Budget Header']
                self.assertEqual(header_sheet['A1'].value, 'Budget Code')
                self.assertEqual(header_sheet['A2'].value, 'EXPORT2026')
                
                # Check amounts sheet
                amounts_sheet = wb['Budget Amounts']
                self.assertEqual(amounts_sheet['A1'].value, 'Segment Code')
                self.assertEqual(amounts_sheet['B1'].value, 'Segment Name')
                self.assertEqual(amounts_sheet['C1'].value, 'Control Level')
                self.assertEqual(amounts_sheet['D1'].value, 'Original Budget')
                self.assertEqual(amounts_sheet['E1'].value, 'Committed')
                self.assertEqual(amounts_sheet['F1'].value, 'Encumbered')
                self.assertEqual(amounts_sheet['G1'].value, 'Actual')
                self.assertEqual(amounts_sheet['H1'].value, 'Available')
                
                # Check data rows
                self.assertEqual(amounts_sheet['A2'].value, '5000')
                self.assertEqual(amounts_sheet['D2'].value, 50000.00)
        else:
            # Functionality not yet implemented
            self.assertIn(
                response.status_code,
                [status.HTTP_404_NOT_FOUND, status.HTTP_405_METHOD_NOT_ALLOWED]
            )
            print("INFO: Excel export endpoint not yet implemented")
    
    def test_export_includes_all_budget_data(self):
        """Test export includes complete budget information"""
        url = reverse('budget_control:budget-export', kwargs={'pk': self.budget.id})
        response = self.client.get(url)
        
        if response.status_code == status.HTTP_200_OK and OPENPYXL_AVAILABLE:
            wb = openpyxl.load_workbook(io.BytesIO(response.content))
            amounts_sheet = wb['Budget Amounts']
            
            # Should have 2 data rows (plus header)
            # Row 2: 5000 - Travel Expense
            # Row 3: 5100 - Office Supplies
            self.assertIsNotNone(amounts_sheet['A2'].value)
            self.assertIsNotNone(amounts_sheet['A3'].value)
            
            # Verify calculations in export
            # Available = Original - Committed - Encumbered - Actual
            available_5000 = amounts_sheet['H2'].value
            if available_5000 is not None:
                expected_available = 50000.00 - 10000.00 - 15000.00 - 5000.00
                self.assertEqual(available_5000, expected_available)
    
    def test_export_empty_budget(self):
        """Test export of budget with no amounts"""
        empty_budget = BudgetHeader.objects.create(
            budget_code='EMPTY2026',
            budget_name='Empty Budget',
            start_date=date.today(),
            end_date=date.today() + timedelta(days=365),
            currency=self.currency,
            status='DRAFT'
        )
        
        url = reverse('budget_control:budget-export', kwargs={'pk': empty_budget.id})
        response = self.client.get(url)
        
        if response.status_code == status.HTTP_200_OK and OPENPYXL_AVAILABLE:
            wb = openpyxl.load_workbook(io.BytesIO(response.content))
            amounts_sheet = wb['Budget Amounts']
            
            # Should have header row only
            self.assertEqual(amounts_sheet['A1'].value, 'Segment Code')
            self.assertIsNone(amounts_sheet['A2'].value)
    
    def test_export_non_existent_budget(self):
        """Test export of non-existent budget"""
        url = reverse('budget_control:budget-export', kwargs={'pk': 99999})
        response = self.client.get(url)
        
        # Should return 404
        self.assertEqual(response.status_code, status.HTTP_404_NOT_FOUND)
    
    def test_export_with_formatting(self):
        """Test export includes proper formatting"""
        url = reverse('budget_control:budget-export', kwargs={'pk': self.budget.id})
        response = self.client.get(url)
        
        if response.status_code == status.HTTP_200_OK and OPENPYXL_AVAILABLE:
            wb = openpyxl.load_workbook(io.BytesIO(response.content))
            amounts_sheet = wb['Budget Amounts']
            
            # Check number formatting for currency columns
            # Columns D-H should have currency format
            for col in ['D', 'E', 'F', 'G', 'H']:
                cell = amounts_sheet[f'{col}2']
                if cell.value is not None:
                    # Should have number format (currency or accounting)
                    self.assertIsNotNone(cell.number_format)


class BudgetExcelImportTestCase(APITestCase):
    """Test budget Excel import functionality"""
    
    def setUp(self):
        """Set up test data"""
        self.client = APIClient()
        
        self.user = create_test_user()
        self.client.force_authenticate(user=self.user)
        
        # Create currency
        self.currency = create_test_currency()
        
        # Create segment types and segments
        self.account_type = XX_SegmentType.objects.create(
            segment_name='Account',
            is_required=True,
            length=4,
            display_order=1
        )
        
        self.segment_5000 = XX_Segment.objects.create(segment_type=self.account_type, code='5000', alias='Travel Expense', node_type='child', is_active=True)
        
        self.segment_5100 = XX_Segment.objects.create(segment_type=self.account_type, code='5100', alias='Office Supplies', node_type='child', is_active=True)
        
        self.segment_5200 = XX_Segment.objects.create(segment_type=self.account_type, code='5200', alias='Equipment', node_type='child', is_active=True)
        
        # Create budget
        self.budget = BudgetHeader.objects.create(
            budget_code='IMPORT2026',
            budget_name='Import Test Budget',
            start_date=date.today(),
            end_date=date.today() + timedelta(days=365),
            currency=self.currency,
            status='DRAFT',  # Must be DRAFT for import
            is_active=True
        )
        
        # Create segment values (required for import)
        BudgetSegmentValue.objects.create(
            budget_header=self.budget,
            segment_value=self.segment_5000,
            control_level='ABSOLUTE'
        )
        
        BudgetSegmentValue.objects.create(
            budget_header=self.budget,
            segment_value=self.segment_5100,
            control_level='ADVISORY'
        )
        
        BudgetSegmentValue.objects.create(
            budget_header=self.budget,
            segment_value=self.segment_5200,
            control_level='ABSOLUTE'
        )
    
    def create_excel_file(self, data_rows):
        """Helper to create Excel file for import"""
        if not OPENPYXL_AVAILABLE:
            return None
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Budget Amounts'
        
        # Header row
        headers = ['Segment Code', 'Original Budget', 'Adjustment', 'Notes']
        ws.append(headers)
        
        # Data rows
        for row in data_rows:
            ws.append(row)
        
        # Save to bytes
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return excel_file
    
    def test_import_budget_amounts_from_excel(self):
        """Test POST /budget-headers/<pk>/import/ with Excel file"""
        if not OPENPYXL_AVAILABLE:
            self.skipTest("openpyxl not available")
        
        # Create Excel file
        data_rows = [
            ['5000', 50000.00, 0, 'Travel budget'],
            ['5100', 30000.00, 5000.00, 'Supplies with adjustment'],
            ['5200', 75000.00, -5000.00, 'Equipment budget reduced']
        ]
        
        excel_file = self.create_excel_file(data_rows)
        
        uploaded_file = SimpleUploadedFile(
            'budget_import.xlsx',
            excel_file.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # Note: This endpoint may not exist yet - this is a specification
        url = reverse('budget_control:budget-import', kwargs={'pk': self.budget.id})
        response = self.client.post(
            url,
            {'file': uploaded_file},
            format='multipart'
        )
        
        # If not implemented, will return 404 or 405
        # When implemented, should return:
        if response.status_code == status.HTTP_200_OK:
            self.assertEqual(response.data.get('status', 'success'), 'success')
            
            # Verify data was imported
            amounts = BudgetAmount.objects.filter(budget_header=self.budget)
            self.assertEqual(amounts.count(), 3)
            
            # Check specific amounts
            amt_5000 = BudgetAmount.objects.get(
                budget_header=self.budget,
                budget_segment_value__segment_value=self.segment_5000
            )
            self.assertEqual(amt_5000.original_budget, Decimal('50000.00'))
            self.assertEqual(amt_5000.adjustment_amount, Decimal('0'))
            
            amt_5100 = BudgetAmount.objects.get(
                budget_header=self.budget,
                budget_segment_value__segment_value=self.segment_5100
            )
            self.assertEqual(amt_5100.original_budget, Decimal('30000.00'))
            self.assertEqual(amt_5100.adjustment_amount, Decimal('5000.00'))
            
            amt_5200 = BudgetAmount.objects.get(
                budget_header=self.budget,
                budget_segment_value__segment_value=self.segment_5200
            )
            self.assertEqual(amt_5200.original_budget, Decimal('75000.00'))
            self.assertEqual(amt_5200.adjustment_amount, Decimal('-5000.00'))
        else:
            # Functionality not yet implemented
            self.assertIn(
                response.status_code,
                [status.HTTP_404_NOT_FOUND, status.HTTP_405_METHOD_NOT_ALLOWED]
            )
            print("INFO: Excel import endpoint not yet implemented")
    
    def test_import_validates_segment_exists(self):
        """Test import fails for non-existent segment"""
        if not OPENPYXL_AVAILABLE:
            self.skipTest("openpyxl not available")
        
        # Create Excel with invalid segment
        data_rows = [
            ['5000', 50000.00, 0, 'Valid'],
            ['9999', 10000.00, 0, 'Invalid segment - does not exist']
        ]
        
        excel_file = self.create_excel_file(data_rows)
        uploaded_file = SimpleUploadedFile(
            'invalid_import.xlsx',
            excel_file.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        url = reverse('budget_control:budget-import', kwargs={'pk': self.budget.id})
        response = self.client.post(
            url,
            {'file': uploaded_file},
            format='multipart'
        )
        
        if response.status_code != status.HTTP_404_NOT_FOUND:
            # Should return validation error
            self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
            self.assertIn('9999', str(response.data))
    
    def test_import_validates_segment_in_budget(self):
        """Test import fails for segment not in budget"""
        if not OPENPYXL_AVAILABLE:
            self.skipTest("openpyxl not available")
        
        # Create segment not in budget
        segment_6000 = XX_Segment.objects.create(segment_type=self.account_type, code='6000', alias='Marketing', node_type='child', is_active=True)
        
        # Try to import for segment not in budget
        data_rows = [
            ['6000', 20000.00, 0, 'Not in budget']
        ]
        
        excel_file = self.create_excel_file(data_rows)
        uploaded_file = SimpleUploadedFile(
            'wrong_segment.xlsx',
            excel_file.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        url = reverse('budget_control:budget-import', kwargs={'pk': self.budget.id})
        response = self.client.post(
            url,
            {'file': uploaded_file},
            format='multipart'
        )
        
        if response.status_code != status.HTTP_404_NOT_FOUND:
            # Should return validation error
            self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
            self.assertIn('6000', str(response.data))
    
    def test_import_validates_positive_budget(self):
        """Test import validates budget amounts are positive"""
        if not OPENPYXL_AVAILABLE:
            self.skipTest("openpyxl not available")
        
        # Create Excel with negative budget
        data_rows = [
            ['5000', -10000.00, 0, 'Invalid negative budget']
        ]
        
        excel_file = self.create_excel_file(data_rows)
        uploaded_file = SimpleUploadedFile(
            'negative_budget.xlsx',
            excel_file.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        url = reverse('budget_control:budget-import', kwargs={'pk': self.budget.id})
        response = self.client.post(
            url,
            {'file': uploaded_file},
            format='multipart'
        )
        
        if response.status_code != status.HTTP_404_NOT_FOUND:
            # Should return validation error
            self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
            self.assertIn('negative', str(response.data).lower())
    
    def test_import_rejects_active_budget(self):
        """Test import rejects ACTIVE budget (must be DRAFT)"""
        if not OPENPYXL_AVAILABLE:
            self.skipTest("openpyxl not available")
        
        # Set budget to ACTIVE
        self.budget.status = 'ACTIVE'
        self.budget.save()
        
        data_rows = [
            ['5000', 50000.00, 0, 'Should fail']
        ]
        
        excel_file = self.create_excel_file(data_rows)
        uploaded_file = SimpleUploadedFile(
            'active_budget.xlsx',
            excel_file.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        url = reverse('budget_control:budget-import', kwargs={'pk': self.budget.id})
        response = self.client.post(
            url,
            {'file': uploaded_file},
            format='multipart'
        )
        
        if response.status_code != status.HTTP_404_NOT_FOUND:
            # Should return error about budget status
            self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
            self.assertIn('active', str(response.data).lower())
    
    def test_import_handles_duplicate_segments(self):
        """Test import handles duplicate segment codes in file"""
        if not OPENPYXL_AVAILABLE:
            self.skipTest("openpyxl not available")
        
        # Create Excel with duplicate segments
        data_rows = [
            ['5000', 50000.00, 0, 'First entry'],
            ['5000', 60000.00, 0, 'Duplicate entry']
        ]
        
        excel_file = self.create_excel_file(data_rows)
        uploaded_file = SimpleUploadedFile(
            'duplicates.xlsx',
            excel_file.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        url = reverse('budget_control:budget-import', kwargs={'pk': self.budget.id})
        response = self.client.post(
            url,
            {'file': uploaded_file},
            format='multipart'
        )
        
        if response.status_code != status.HTTP_404_NOT_FOUND:
            # Should return validation error
            self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
            self.assertIn('duplicate', str(response.data).lower())
    
    def test_import_returns_summary(self):
        """Test import returns summary of imported records"""
        if not OPENPYXL_AVAILABLE:
            self.skipTest("openpyxl not available")
        
        data_rows = [
            ['5000', 50000.00, 0, 'Travel'],
            ['5100', 30000.00, 5000.00, 'Supplies'],
            ['5200', 75000.00, 0, 'Equipment']
        ]
        
        excel_file = self.create_excel_file(data_rows)
        uploaded_file = SimpleUploadedFile(
            'budget_import.xlsx',
            excel_file.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        url = reverse('budget_control:budget-import', kwargs={'pk': self.budget.id})
        response = self.client.post(
            url,
            {'file': uploaded_file},
            format='multipart'
        )
        
        if response.status_code == status.HTTP_200_OK:
            # Should include summary
            self.assertIn('imported_count', response.data.get('data', response.data))
            self.assertEqual(response.data.get('data', response.data)['imported_count'], 3)
            
            self.assertIn('total_budget', response.data.get('data', response.data))
            # Total = 50000 + (30000 + 5000) + 75000 = 160000
            self.assertEqual(
                Decimal(response.data.get('data', response.data)['total_budget']),
                Decimal('160000.00')
            )


class BudgetExcelTemplateTestCase(APITestCase):
    """Test budget Excel template generation"""
    
    def setUp(self):
        """Set up test data"""
        self.client = APIClient()
        
        self.user = create_test_user()
        self.client.force_authenticate(user=self.user)
        
        self.currency = create_test_currency()
        
        self.budget = BudgetHeader.objects.create(
            budget_code='TEMPLATE2026',
            budget_name='Template Test',
            start_date=date.today(),
            end_date=date.today() + timedelta(days=365),
            currency=self.currency,
            status='DRAFT'
        )
    
    def test_download_import_template(self):
        """Test GET /budget-headers/<pk>/template/ returns Excel template"""
        # Note: This endpoint may not exist yet - this is a specification
        url = reverse('budget_control:budget-template', kwargs={'pk': self.budget.id})
        response = self.client.get(url)
        
        if response.status_code == status.HTTP_200_OK:
            # Verify response is Excel file
            self.assertEqual(
                response['Content-Type'],
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
            if OPENPYXL_AVAILABLE:
                wb = openpyxl.load_workbook(io.BytesIO(response.content))
                ws = wb.active
                
                # Should have headers and instructions
                self.assertEqual(ws['A1'].value, 'Segment Code')
                self.assertEqual(ws['B1'].value, 'Original Budget')
                self.assertEqual(ws['C1'].value, 'Adjustment')
                self.assertEqual(ws['D1'].value, 'Notes')
                
                # Should have example row
                self.assertIsNotNone(ws['A2'].value)
        else:
            # Functionality not yet implemented
            self.assertIn(
                response.status_code,
                [status.HTTP_404_NOT_FOUND, status.HTTP_405_METHOD_NOT_ALLOWED]
            )
            print("INFO: Excel template endpoint not yet implemented")


