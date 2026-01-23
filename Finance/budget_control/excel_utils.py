"""
Excel Import/Export Utilities for Budget Control
"""
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from django.http import HttpResponse
from decimal import Decimal, InvalidOperation
from io import BytesIO


def export_budget_to_excel(budget):
    """
    Export a budget header with all amounts to Excel file
    
    Args:
        budget: BudgetHeader instance
    
    Returns:
        HttpResponse with Excel file
    """
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create Budget Header sheet
    ws_header = wb.create_sheet('Budget Header')
    
    # Budget Header - Row 1 has headers, Row 2 has values
    header_labels = ['Budget Code', 'Budget Name', 'Description', 'Status', 'Start Date', 'End Date', 'Currency', 'Default Control Level', 'Is Active']
    header_values = [
        budget.budget_code,
        budget.budget_name,
        budget.description or '',
        budget.status,
        budget.start_date.strftime('%Y-%m-%d'),
        budget.end_date.strftime('%Y-%m-%d'),
        budget.currency.code,
        budget.default_control_level,
        'Yes' if budget.is_active else 'No'
    ]
    
    # Row 1: Labels
    for col_idx, label in enumerate(header_labels, start=1):
        cell = ws_header.cell(row=1, column=col_idx, value=label)
        cell.font = Font(bold=True)
    
    # Row 2: Values
    for col_idx, value in enumerate(header_values, start=1):
        ws_header.cell(row=2, column=col_idx, value=value)
    
    # Auto-fit columns
    for col_idx in range(1, len(header_labels) + 1):
        col_letter = chr(64 + col_idx)
        ws_header.column_dimensions[col_letter].width = 20
    
    # Create Budget Amounts sheet
    ws_amounts = wb.create_sheet('Budget Amounts')
    
    # Headers - Simplified based on test expectations
    headers = [
        'Segment Code',
        'Segment Name',
        'Control Level',
        'Original Budget',
        'Committed',
        'Encumbered',
        'Actual',
        'Available'
    ]
    
    # Add headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_amounts.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
    
    # Data rows
    amounts = budget.budget_amounts.select_related(
        'budget_segment_value__segment_value__segment_type'
    ).all()
    
    for row_idx, amount in enumerate(amounts, start=2):
        segment = amount.budget_segment_value.segment_value
        
        data_row = [
            segment.code,
            segment.alias,
            amount.get_effective_control_level(),
            float(amount.original_budget),
            float(amount.committed_amount),
            float(amount.encumbered_amount),
            float(amount.actual_amount),
            float(amount.get_available()),
        ]
        
        for col_idx, value in enumerate(data_row, start=1):
            ws_amounts.cell(row=row_idx, column=col_idx, value=value)
    
    # Auto-fit columns
    for col in range(1, len(headers) + 1):
        col_letter = chr(64 + col)
        ws_amounts.column_dimensions[col_letter].width = 15
    
    # Prepare response
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=Budget_{budget.budget_code}.xlsx'
    
    return response


def create_budget_template(budget):
    """
    Create an Excel template for importing budget amounts
    
    Args:
        budget: BudgetHeader instance
    
    Returns:
        HttpResponse with Excel template file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Budget Import Template'
    
    # Headers - Simplified based on test expectations
    headers = [
        'Segment Code',
        'Original Budget',
        'Adjustment',
        'Notes'
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        col_letter = chr(64 + col_idx)
        ws.column_dimensions[col_letter].width = 20
    
    # Add instruction/example row
    ws.cell(row=2, column=1, value='5000')
    ws.cell(row=2, column=2, value=50000.00)
    ws.cell(row=2, column=3, value=0.00)
    ws.cell(row=2, column=4, value='Example budget entry')
    
    # Prepare response
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=Budget_Import_Template_{budget.budget_code}.xlsx'
    
    return response


def import_budget_from_excel(budget, excel_file):
    """
    Import budget amounts from Excel file
    
    Args:
        budget: BudgetHeader instance
        excel_file: Uploaded Excel file
    
    Returns:
        dict with import results
    """
    from Finance.GL.models import XX_Segment
    from Finance.budget_control.models import BudgetAmount, BudgetSegmentValue
    
    # Check if budget is DRAFT - only DRAFT budgets can be imported
    if budget.status != 'DRAFT':
        return {
            'error': 'Only DRAFT budgets can be imported. Current status: ' + budget.status,
            'total_rows': 0,
            'success_count': 0,
            'error_count': 0,
            'errors': []
        }
    
    wb = load_workbook(excel_file)
    ws = wb.active
    
    results = {
        'total_rows': 0,
        'success_count': 0,
        'imported_count': 0,  # Add this field for test compatibility
        'error_count': 0,
        'errors': [],
        'total_budget': Decimal('0')  # Track total budget
    }
    
    # Track seen segment codes to detect duplicates
    seen_segments = set()
    
    # Skip header row
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    
    for row_idx, row in enumerate(rows, start=2):
        # Skip empty rows
        if not any(row):
            continue
            
        results['total_rows'] += 1
        
        try:
            # Extract data from row: Segment Code, Original Budget, Adjustment, Notes
            segment_code = str(row[0]).strip() if row[0] else None
            original_budget = row[1] if row[1] is not None else Decimal('0')
            adjustment_amount = row[2] if row[2] is not None else Decimal('0')
            notes = str(row[3]).strip() if row[3] else ''
            
            # Validate required fields
            if not segment_code:
                results['errors'].append(f"Row {row_idx}: Missing segment code")
                results['error_count'] += 1
                continue
            
            # Check for duplicate segments in the file
            if segment_code in seen_segments:
                results['errors'].append(f"Row {row_idx}: Duplicate segment code '{segment_code}' in file")
                results['error_count'] += 1
                continue
            
            seen_segments.add(segment_code)
            
            # Convert to Decimal
            try:
                original_budget = Decimal(str(original_budget))
                adjustment_amount = Decimal(str(adjustment_amount))
            except (InvalidOperation, ValueError) as e:
                results['errors'].append(f"Row {row_idx}: Invalid number format")
                results['error_count'] += 1
                continue
            
            # Validate positive budget
            if original_budget < 0:
                results['errors'].append(f"Row {row_idx}: Original budget must be non-negative")
                results['error_count'] += 1
                continue
            
            # Find segment from active segments in this budget
            existing_seg_vals = budget.budget_segment_values.select_related('segment_value').all()
            matching_seg_val = None
            
            for seg_val in existing_seg_vals:
                if seg_val.segment_value.code == segment_code:
                    matching_seg_val = seg_val
                    break
            
            if not matching_seg_val:
                results['errors'].append(
                    f"Row {row_idx}: Segment {segment_code} not found in budget"
                )
                results['error_count'] += 1
                continue
            
            # Create or update budget amount
            amount, created = BudgetAmount.objects.update_or_create(
                budget_header=budget,
                budget_segment_value=matching_seg_val,
                defaults={
                    'original_budget': original_budget,
                    'adjustment_amount': adjustment_amount,
                    'notes': notes
                }
            )
            
            results['success_count'] += 1
            results['imported_count'] += 1  # Track imported count
            
            # Add to total budget (original + adjustment)
            results['total_budget'] += original_budget + adjustment_amount
            
        except Exception as e:
            results['errors'].append(f"Row {row_idx}: {str(e)}")
            results['error_count'] += 1
    
    # Convert total_budget to string for JSON serialization
    results['total_budget'] = str(results['total_budget'])
    
    return results
