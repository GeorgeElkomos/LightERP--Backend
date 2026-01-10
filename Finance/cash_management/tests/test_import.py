"""
Tests for Bank Statement Import Functionality

This module tests the Excel/CSV import feature for bank statements including:
- File upload validation
- Column mapping and data parsing
- Preview functionality (validation without saving)
- Full import with database transactions
- Error handling for invalid files
- Template download
"""

import io
import pytest
from decimal import Decimal
from datetime import date
from django.core.files.uploadedfile import SimpleUploadedFile
from rest_framework.test import APITestCase
from rest_framework import status
from django.contrib.auth import get_user_model
from openpyxl import Workbook

from Finance.core.models import Currency, Country
from Finance.cash_management.models import Bank, BankBranch, BankAccount, BankStatement, BankStatementLine
from Finance.cash_management.services.statement_import import BankStatementImporter
from Finance.GL.models import XX_SegmentType, XX_Segment, XX_Segment_combination

User = get_user_model()


class BankStatementImportTestCase(APITestCase):
    """Test cases for bank statement import functionality"""

    def setUp(self):
        """Set up test data"""
        # Create user
        self.user = User.objects.create_user(
            email='testuser@example.com',
            name='Test User',
            phone_number='1234567890',
            password='testpass123'
        )
        self.client.force_authenticate(user=self.user)

        # Create country
        self.country = Country.objects.create(
            name='United States',
            code='US'
        )

        # Create currency
        self.currency = Currency.objects.create(
            code='USD',
            name='US Dollar',
            symbol='$'
        )

        # Create bank first
        self.bank = Bank.objects.create(
            bank_name='Test Bank',
            bank_code='TB01',
            swift_code='TESTUS33',
            country=self.country,
            is_active=True,
            created_by=self.user,
            updated_by=self.user
        )

        # Create branch with bank
        self.branch = BankBranch.objects.create(
            bank=self.bank,
            branch_name='Main Branch',
            branch_code='MB01',
            country=self.country,
            is_active=True,
            created_by=self.user,
            updated_by=self.user
        )

        # Create GL segments for bank account
        self.segment_type = XX_SegmentType.objects.create(
            segment_name='Department',
            is_required=True,
            display_order=1,
            is_active=True
        )
        
        self.segment = XX_Segment.objects.create(
            segment_type=self.segment_type,
            code='CASH',
            alias='Cash Department',
            node_type='child',
            is_active=True
        )
        
        self.cash_combo = XX_Segment_combination.get_combination_id(
            [(self.segment_type.id, 'CASH')],
            'Cash Account'
        )
        
        self.clearing_combo = XX_Segment_combination.get_combination_id(
            [(self.segment_type.id, 'CASH')],
            'Cash Clearing'
        )

        # Create bank account
        self.bank_account = BankAccount.objects.create(
            account_number='1234567890',
            account_name='Test Account',
            branch=self.branch,
            currency=self.currency,
            account_type='CHECKING',
            is_active=True,
            cash_GL_combination_id=self.cash_combo,
            cash_clearing_GL_combination_id=self.clearing_combo,
            created_by=self.user
        )

    def create_csv_file(self, content):
        """Helper to create CSV file"""
        return SimpleUploadedFile(
            'test_statement.csv',
            content.encode('utf-8'),
            content_type='text/csv'
        )

    def create_excel_file(self, data):
        """Helper to create Excel file"""
        wb = Workbook()
        ws = wb.active
        ws.title = 'Statement Data'

        # Write headers
        headers = list(data[0].keys())
        ws.append(headers)

        # Write data rows
        for row in data:
            ws.append(list(row.values()))

        # Save to bytes
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return SimpleUploadedFile(
            'test_statement.xlsx',
            excel_file.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    def test_import_csv_success(self):
        """Test successful CSV import"""
        csv_content = """Transaction Date,Description,Reference,Debit,Credit,Balance
2024-01-01,Opening Balance,,,,1000.00
2024-01-02,Deposit from Customer,REF001,,500.00,1500.00
2024-01-03,Payment to Supplier,REF002,200.00,,1300.00
2024-01-04,Bank Charges,REF003,10.00,,1290.00
"""
        
        csv_file = self.create_csv_file(csv_content)

        data = {
            'file': csv_file,
            'bank_account_id': self.bank_account.id,
            'statement_number': 'STMT-2024-001',
            'statement_date': '2024-01-04',
            'opening_balance': '1000.00',
            'closing_balance': '1290.00'
        }

        response = self.client.post(
            '/finance/cash/statements/import_statement/',
            data,
            format='multipart'
        )

        # Debug output
        if response.status_code != status.HTTP_201_CREATED:
            print(f"\nError Response: {response.status_code}")
            print(f"Response data: {response.data}")
            if 'errors' in response.data and response.data['errors']:
                print(f"Errors: {response.data['errors']}")

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertTrue(response.data['success'])
        self.assertEqual(response.data['lines_created'], 3)  # Excluding opening balance

        # Verify database records
        statement = BankStatement.objects.get(statement_number='STMT-2024-001')
        self.assertEqual(statement.bank_account, self.bank_account)
        self.assertEqual(statement.opening_balance, Decimal('1000.00'))
        self.assertEqual(statement.closing_balance, Decimal('1290.00'))
        self.assertEqual(statement.statement_lines.count(), 3)

    def test_import_excel_success(self):
        """Test successful Excel import"""
        excel_data = [
            {
                'Transaction Date': '2024-01-02',
                'Description': 'Customer Payment',
                'Reference': 'REF001',
                'Debit': '',
                'Credit': '500.00',
                'Balance': '1500.00'
            }
        ]

        excel_file = self.create_excel_file(excel_data)

        data = {
            'file': excel_file,
            'bank_account_id': self.bank_account.id,
            'statement_number': 'STMT-2024-002',
            'statement_date': '2024-01-02',
            'opening_balance': '1000.00',
            'closing_balance': '1500.00'
        }

        response = self.client.post(
            '/finance/cash/statements/import_statement/',
            data,
            format='multipart'
        )

        if response.status_code != status.HTTP_201_CREATED:
            print(f"\nExcel Error: {response.status_code}")
            print(f"Response: {response.data}")

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertTrue(response.data['success'])
        self.assertEqual(response.data['lines_created'], 1)

    def test_import_preview_success(self):
        """Test preview without saving to database"""
        csv_content = """Transaction Date,Description,Debit,Credit
2024-01-02,Test Transaction 1,100.00,
2024-01-03,Test Transaction 2,,200.00
2024-01-04,Test Transaction 3,50.00,
"""
        
        csv_file = self.create_csv_file(csv_content)

        data = {
            'file': csv_file,
            'bank_account_id': self.bank_account.id
        }

        # Count statements before
        initial_count = BankStatement.objects.count()

        response = self.client.post(
            '/finance/cash/statements/import_preview/',
            data,
            format='multipart'
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['success'])
        self.assertIn('lines_preview', response.data)
        self.assertIn('summary', response.data)
        
        # Verify no database records created
        self.assertEqual(BankStatement.objects.count(), initial_count)

    def test_import_invalid_file_format(self):
        """Test rejection of invalid file format"""
        # Create a text file
        invalid_file = SimpleUploadedFile(
            'test.txt',
            b'This is not a valid CSV or Excel file',
            content_type='text/plain'
        )

        data = {
            'file': invalid_file,
            'bank_account_id': self.bank_account.id,
            'statement_number': 'STMT-2024-003',
            'statement_date': '2024-01-01'
        }

        response = self.client.post(
            '/finance/cash/statements/import_statement/',
            data,
            format='multipart'
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_import_file_too_large(self):
        """Test rejection of files exceeding size limit"""
        # Create a large CSV content (>10MB)
        large_content = "col1,col2,col3\n" * 500000  # Creates large file
        large_file = SimpleUploadedFile(
            'large_file.csv',
            large_content.encode('utf-8'),
            content_type='text/csv'
        )

        data = {
            'file': large_file,
            'bank_account_id': self.bank_account.id,
            'statement_number': 'STMT-2024-004',
            'statement_date': '2024-01-01'
        }

        response = self.client.post(
            '/finance/cash/statements/import_statement/',
            data,
            format='multipart'
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_import_missing_required_columns(self):
        """Test error handling for missing required columns"""
        csv_content = """Date,Amount
2024-01-01,100.00
"""
        
        csv_file = self.create_csv_file(csv_content)

        data = {
            'file': csv_file,
            'bank_account_id': self.bank_account.id,
            'statement_number': 'STMT-2024-005',
            'statement_date': '2024-01-01'
        }

        response = self.client.post(
            '/finance/cash/statements/import_statement/',
            data,
            format='multipart'
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertFalse(response.data['success'])

    def test_import_invalid_dates(self):
        """Test error handling for invalid date formats"""
        csv_content = """Transaction Date,Description,Debit,Credit
invalid-date,Test Transaction,100.00,
2024-13-45,Test Transaction 2,,200.00
"""
        
        csv_file = self.create_csv_file(csv_content)

        data = {
            'file': csv_file,
            'bank_account_id': self.bank_account.id,
            'statement_number': 'STMT-2024-006',
            'statement_date': '2024-01-01'
        }

        response = self.client.post(
            '/finance/cash/statements/import_statement/',
            data,
            format='multipart'
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertFalse(response.data['success'])
        self.assertIn('errors', response.data)

    def test_import_duplicate_statement_number(self):
        """Test rejection of duplicate statement numbers"""
        # Create existing statement
        BankStatement.objects.create(
            bank_account=self.bank_account,
            statement_number='STMT-DUPLICATE',
            statement_date=date(2024, 1, 1),
            from_date=date(2024, 1, 1),
            to_date=date(2024, 1, 1),
            opening_balance=Decimal('1000.00'),
            closing_balance=Decimal('1000.00'),
            created_by=self.user,
            updated_by=self.user
        )

        csv_content = """Transaction Date,Description,Debit,Credit
2024-01-02,Test Transaction,100.00,
"""
        
        csv_file = self.create_csv_file(csv_content)

        data = {
            'file': csv_file,
            'bank_account_id': self.bank_account.id,
            'statement_number': 'STMT-DUPLICATE',
            'statement_date': '2024-01-02'
        }

        response = self.client.post(
            '/finance/cash/statements/import_statement/',
            data,
            format='multipart'
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_import_invalid_bank_account(self):
        """Test error when bank account doesn't exist"""
        csv_content = """Transaction Date,Description,Debit,Credit
2024-01-01,Test Transaction,100.00,
"""
        
        csv_file = self.create_csv_file(csv_content)

        data = {
            'file': csv_file,
            'bank_account_id': 99999,  # Non-existent ID
            'statement_number': 'STMT-2024-007',
            'statement_date': '2024-01-01'
        }

        response = self.client.post(
            '/finance/cash/statements/import_statement/',
            data,
            format='multipart'
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_download_template(self):
        """Test Excel template download"""
        response = self.client.get(
            '/finance/cash/statements/download_template/'
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        self.assertIn('attachment', response['Content-Disposition'])
        self.assertIn('.xlsx', response['Content-Disposition'])

    def test_import_with_balance_mismatch(self):
        """Test rejection when calculated balance doesn't match closing balance"""
        csv_content = """Transaction Date,Description,Debit,Credit,Balance
2024-01-02,Test Transaction,100.00,,900.00
"""
        
        csv_file = self.create_csv_file(csv_content)

        data = {
            'file': csv_file,
            'bank_account_id': self.bank_account.id,
            'statement_number': 'STMT-2024-008',
            'statement_date': '2024-01-02',
            'opening_balance': '1000.00',
            'closing_balance': '1200.00'  # Mismatch with calculated 900.00
        }

        response = self.client.post(
            '/finance/cash/statements/import_statement/',
            data,
            format='multipart'
        )

        # Should reject due to balance mismatch
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertFalse(response.data['success'])
        self.assertIn('balance mismatch', response.data['message'].lower())

    def test_import_unauthenticated(self):
        """Test import requires authentication"""
        self.client.force_authenticate(user=None)

        csv_content = """Transaction Date,Description,Debit,Credit
2024-01-01,Test,100.00,
"""
        
        csv_file = self.create_csv_file(csv_content)

        data = {
            'file': csv_file,
            'bank_account_id': self.bank_account.id,
            'statement_number': 'STMT-2024-009',
            'statement_date': '2024-01-01'
        }

        response = self.client.post(
            '/finance/cash/statements/import_statement/',
            data,
            format='multipart'
        )

        self.assertEqual(response.status_code, status.HTTP_401_UNAUTHORIZED)


class BankStatementImporterUnitTestCase(APITestCase):
    """Unit tests for BankStatementImporter service class"""

    def setUp(self):
        """Set up test data"""
        self.user = User.objects.create_user(
            email='testuser@example.com',
            name='Test User',
            phone_number='1234567890',
            password='testpass123'
        )

        self.country = Country.objects.create(
            name='United States',
            code='US'
        )

        self.currency = Currency.objects.create(
            code='USD',
            name='US Dollar',
            symbol='$'
        )

        self.bank = Bank.objects.create(
            bank_name='Test Bank',
            bank_code='TB01',
            country=self.country,
            is_active=True,
            created_by=self.user,
            updated_by=self.user
        )

        self.branch = BankBranch.objects.create(
            bank=self.bank,
            branch_name='Main Branch',
            branch_code='MB01',
            country=self.country,
            is_active=True,
            created_by=self.user,
            updated_by=self.user
        )

        # Create GL segments for bank account
        self.segment_type = XX_SegmentType.objects.create(
            segment_name='Department',
            is_required=True,
            display_order=1,
            is_active=True
        )
        
        self.segment = XX_Segment.objects.create(
            segment_type=self.segment_type,
            code='CASH',
            alias='Cash Department',
            node_type='child',
            is_active=True
        )
        
        self.cash_combo = XX_Segment_combination.get_combination_id(
            [(self.segment_type.id, 'CASH')],
            'Cash Account'
        )
        
        self.clearing_combo = XX_Segment_combination.get_combination_id(
            [(self.segment_type.id, 'CASH')],
            'Cash Clearing'
        )

        self.bank_account = BankAccount.objects.create(
            account_number='1234567890',
            account_name='Test Account',
            branch=self.branch,
            currency=self.currency,
            account_type='CHECKING',
            is_active=True,
            cash_GL_combination_id=self.cash_combo,
            cash_clearing_GL_combination_id=self.clearing_combo,
            created_by=self.user
        )

    def test_column_mapping_flexibility(self):
        """Test that importer handles various column name variations"""
        # Test with different column name variations
        test_variations = [
            'transaction_date',
            'trans_date',
            'date',
            'value_date',
            'posting_date'
        ]

        importer = BankStatementImporter(
            file_obj=None,
            bank_account_id=self.bank_account.id,
            user=self.user
        )

        for variation in test_variations:
            # Create mock dataframe with this column name
            import pandas as pd
            df = pd.DataFrame({
                variation: ['2024-01-01'],
                'description': ['Test'],
                'debit': [100.00],
                'credit': [0]
            })
            
            importer.df = df
            result = importer._find_column('transaction_date')
            self.assertEqual(result, variation)

    def test_date_format_parsing(self):
        """Test multiple date format parsing"""
        importer = BankStatementImporter(
            file_obj=None,
            bank_account_id=self.bank_account.id,
            user=self.user
        )

        # Test various date formats
        date_formats = [
            ('2024-01-15', date(2024, 1, 15)),
            ('01/15/2024', date(2024, 1, 15)),
            ('15/01/2024', date(2024, 1, 15)),
            ('15-01-2024', date(2024, 1, 15)),
            ('2024.01.15', date(2024, 1, 15)),
        ]

        for date_str, expected in date_formats:
            result = importer._parse_date(date_str)
            self.assertEqual(result, expected, f"Failed to parse {date_str}")

    def test_transaction_type_detection(self):
        """Test automatic transaction type detection"""
        importer = BankStatementImporter(
            file_obj=None,
            bank_account_id=self.bank_account.id,
            user=self.user
        )

        # Test with debit/credit columns
        row1 = {'debit_amount': Decimal('100.00'), 'credit_amount': Decimal('0')}
        trans_type, amount = importer._determine_transaction_type(row1)
        self.assertEqual(trans_type, 'DEBIT')
        self.assertEqual(amount, Decimal('100.00'))

        row2 = {'debit_amount': Decimal('0'), 'credit_amount': Decimal('200.00')}
        trans_type, amount = importer._determine_transaction_type(row2)
        self.assertEqual(trans_type, 'CREDIT')
        self.assertEqual(amount, Decimal('200.00'))

    def test_decimal_parsing_with_currency_symbols(self):
        """Test parsing of amounts with currency symbols"""
        importer = BankStatementImporter(
            file_obj=None,
            bank_account_id=self.bank_account.id,
            user=self.user
        )

        test_values = [
            ('$1,234.56', Decimal('1234.56')),
            ('€1.234,56', Decimal('1234.56')),
            ('(1,000.00)', Decimal('-1000.00')),  # Negative in parentheses
            ('1234.56', Decimal('1234.56')),
        ]

        for value_str, expected in test_values:
            result = importer._parse_decimal(value_str)
            self.assertEqual(result, expected, f"Failed to parse {value_str}")


@pytest.mark.django_db
class TestBankStatementImportIntegration:
    """Integration tests using pytest"""

    def test_full_import_workflow(self, client, django_user_model):
        """Test complete import workflow from upload to database"""
        # Setup
        user = django_user_model.objects.create_user(
            email='testuser@example.com',
            name='Test User',
            phone_number='1234567890',
            password='testpass'
        )
        client.force_authenticate(user=user)

        # Create required objects
        country = Country.objects.create(
            name='United States',
            code='US'
        )
        currency = Currency.objects.create(
            code='USD',
            name='US Dollar',
            symbol='$'
        )
        bank = Bank.objects.create(
            bank_name='Test Bank',
            bank_code='TB01',
            country=country,
            is_active=True,
            created_by=user,
            updated_by=user
        )
        branch = BankBranch.objects.create(
            bank=bank,
            branch_name='Test Branch',
            branch_code='TB01',
            country=country,
            is_active=True,
            created_by=user,
            updated_by=user
        )
        
        # Create GL segments
        segment_type = XX_SegmentType.objects.create(
            segment_name='Department',
            is_required=True,
            display_order=1,
            is_active=True
        )
        
        segment = XX_Segment.objects.create(
            segment_type=segment_type,
            code='CASH',
            alias='Cash Department',
            node_type='child',
            is_active=True
        )
        
        cash_combo = XX_Segment_combination.get_combination_id(
            [(segment_type.id, 'CASH')],
            'Cash Account'
        )
        
        clearing_combo = XX_Segment_combination.get_combination_id(
            [(segment_type.id, 'CASH')],
            'Cash Clearing'
        )
        
        bank_account = BankAccount.objects.create(
            account_number='123456',
            account_name='Test Account',
            branch=branch,
            currency=currency,
            account_type='CHECKING',
            is_active=True,
            cash_GL_combination_id=cash_combo,
            cash_clearing_GL_combination_id=clearing_combo,
            created_by=user
        )

        # Create test file
        csv_content = """Transaction Date,Description,Reference,Debit,Credit
2024-01-02,Customer Payment,REF001,,1000.00
2024-01-03,Supplier Payment,REF002,500.00,
"""
        csv_file = SimpleUploadedFile(
            'test.csv',
            csv_content.encode('utf-8'),
            content_type='text/csv'
        )

        # Import
        response = client.post(
            '/finance/cash/statements/import_statement/',
            {
                'file': csv_file,
                'bank_account_id': bank_account.id,
                'statement_number': 'STMT-INT-001',
                'statement_date': '2024-01-03',
                'opening_balance': '0.00',
                'closing_balance': '500.00'
            },
            format='multipart'
        )

        assert response.status_code == 201
        assert BankStatement.objects.filter(statement_number='STMT-INT-001').exists()
        statement = BankStatement.objects.get(statement_number='STMT-INT-001')
        assert statement.statement_lines.count() == 2


class BankStatementImporterServiceTestCase(APITestCase):
    """Additional tests for BankStatementImporter service methods"""

    def setUp(self):
        """Set up test data"""
        self.user = User.objects.create_user(
            email='servicetest@example.com',
            name='Service Test User',
            phone_number='9876543210',
            password='testpass123'
        )

        self.country = Country.objects.create(
            name='United Kingdom',
            code='GB'
        )

        self.currency = Currency.objects.create(
            code='GBP',
            name='British Pound',
            symbol='£'
        )

        self.bank = Bank.objects.create(
            bank_name='Service Test Bank',
            bank_code='STB01',
            swift_code='TSTGB33',
            country=self.country,
            is_active=True,
            created_by=self.user,
            updated_by=self.user
        )

        self.branch = BankBranch.objects.create(
            bank=self.bank,
            branch_name='Service Test Branch',
            branch_code='STB01',
            country=self.country,
            is_active=True,
            created_by=self.user,
            updated_by=self.user
        )

        self.segment_type = XX_SegmentType.objects.create(
            segment_name='Service Segment',
            is_required=True,
            display_order=1,
            is_active=True
        )
        
        self.segment = XX_Segment.objects.create(
            segment_type=self.segment_type,
            code='SRV',
            alias='Service',
            node_type='child',
            is_active=True
        )
        
        self.cash_combo = XX_Segment_combination.get_combination_id(
            [(self.segment_type.id, 'SRV')],
            'Service Cash'
        )
        
        self.clearing_combo = XX_Segment_combination.get_combination_id(
            [(self.segment_type.id, 'SRV')],
            'Service Clearing'
        )

        self.bank_account = BankAccount.objects.create(
            account_number='SRV123456',
            account_name='Service Test Account',
            branch=self.branch,
            currency=self.currency,
            account_type='SAVINGS',
            is_active=True,
            cash_GL_combination_id=self.cash_combo,
            cash_clearing_GL_combination_id=self.clearing_combo,
            created_by=self.user
        )

    def test_importer_read_csv_file(self):
        """Test that importer can read and parse CSV files"""
        csv_content = """Transaction Date,Description,Debit,Credit,Balance
2024-02-01,Test Transaction,100.00,,900.00
"""
        csv_file = SimpleUploadedFile(
            'test.csv',
            csv_content.encode('utf-8'),
            content_type='text/csv'
        )

        importer = BankStatementImporter(
            file_obj=csv_file,
            bank_account_id=self.bank_account.id,
            user=self.user
        )

        result = importer.read_file()
        self.assertTrue(result)
        self.assertIsNotNone(importer.df)
        self.assertEqual(len(importer.df), 1)

    def test_importer_read_excel_file(self):
        """Test that importer can read and parse Excel files"""
        # Create Excel file
        wb = Workbook()
        ws = wb.active
        ws.append(['Transaction Date', 'Description', 'Debit', 'Credit', 'Balance'])
        ws.append(['2024-02-01', 'Test Transaction', 100.00, '', 900.00])
        
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        file = SimpleUploadedFile(
            'test.xlsx',
            excel_file.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        importer = BankStatementImporter(
            file_obj=file,
            bank_account_id=self.bank_account.id,
            user=self.user
        )

        result = importer.read_file()
        self.assertTrue(result)
        self.assertIsNotNone(importer.df)
        self.assertEqual(len(importer.df), 1)

    def test_importer_parse_data_extracts_lines(self):
        """Test that parse_data correctly extracts transaction lines"""
        csv_content = """Transaction Date,Description,Debit,Credit,Balance
2024-02-01,Payment Received,,500.00,1500.00
2024-02-02,Payment Sent,200.00,,1300.00
"""
        csv_file = SimpleUploadedFile(
            'test.csv',
            csv_content.encode('utf-8'),
            content_type='text/csv'
        )

        importer = BankStatementImporter(
            file_obj=csv_file,
            bank_account_id=self.bank_account.id,
            user=self.user
        )

        importer.read_file()
        parsed_data = importer.parse_data()

        self.assertEqual(len(parsed_data['lines']), 2)
        self.assertEqual(parsed_data['summary']['valid_lines'], 2)
        self.assertEqual(parsed_data['summary']['error_lines'], 0)
        
        # Check first line (credit)
        self.assertEqual(parsed_data['lines'][0]['transaction_type'], 'CREDIT')
        self.assertEqual(parsed_data['lines'][0]['amount'], Decimal('500.00'))
        
        # Check second line (debit)
        self.assertEqual(parsed_data['lines'][1]['transaction_type'], 'DEBIT')
        self.assertEqual(parsed_data['lines'][1]['amount'], Decimal('200.00'))

    def test_importer_preview_does_not_save(self):
        """Test that preview_import doesn't create database records"""
        csv_content = """Transaction Date,Description,Debit,Credit,Balance
2024-02-01,Preview Test,,100.00,1100.00
"""
        csv_file = SimpleUploadedFile(
            'preview_test.csv',
            csv_content.encode('utf-8'),
            content_type='text/csv'
        )

        initial_statement_count = BankStatement.objects.count()
        initial_line_count = BankStatementLine.objects.count()

        importer = BankStatementImporter(
            file_obj=csv_file,
            bank_account_id=self.bank_account.id,
            user=self.user
        )

        preview_result = importer.preview_import()

        # Verify no new records created
        self.assertEqual(BankStatement.objects.count(), initial_statement_count)
        self.assertEqual(BankStatementLine.objects.count(), initial_line_count)
        
        # Verify preview data returned
        self.assertTrue(preview_result['success'])
        self.assertEqual(preview_result['total_lines'], 1)

    def test_template_contains_required_columns(self):
        """Test that downloaded template has all required columns"""
        self.client.force_authenticate(user=self.user)
        
        response = self.client.get('/finance/cash/statements/download_template/')
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        
        # Parse the response content as Excel
        excel_content = io.BytesIO(response.content)
        from openpyxl import load_workbook
        wb = load_workbook(excel_content)
        
        # Check that both sheets exist
        self.assertIn('Instructions', wb.sheetnames)
        self.assertIn('Statement Data', wb.sheetnames)
        
        # Get data sheet
        ws = wb['Statement Data']
        
        # Get header row (first row)
        headers = [cell.value for cell in ws[1]]
        
        # Verify required columns exist
        required_columns = [
            'Transaction Date',
            'Description',
            'Debit Amount',
            'Credit Amount',
        ]
        
        for col in required_columns:
            self.assertIn(col, headers, f"Missing required column: {col}")

    def test_importer_handles_empty_file(self):
        """Test that importer handles empty files gracefully"""
        csv_content = """Transaction Date,Description,Debit,Credit,Balance
"""
        csv_file = SimpleUploadedFile(
            'empty.csv',
            csv_content.encode('utf-8'),
            content_type='text/csv'
        )

        importer = BankStatementImporter(
            file_obj=csv_file,
            bank_account_id=self.bank_account.id,
            user=self.user
        )

        importer.read_file()
        parsed_data = importer.parse_data()

        self.assertEqual(len(parsed_data['lines']), 0)
        self.assertEqual(parsed_data['summary']['total_lines'], 0)

    def test_importer_handles_special_characters_in_description(self):
        """Test that importer handles special characters in description"""
        csv_content = """Transaction Date,Description,Debit,Credit,Balance
2024-02-01,"Payment for Invoice #123 - Café & Co.",100.00,,900.00
"""
        csv_file = SimpleUploadedFile(
            'special_chars.csv',
            csv_content.encode('utf-8'),
            content_type='text/csv'
        )

        importer = BankStatementImporter(
            file_obj=csv_file,
            bank_account_id=self.bank_account.id,
            user=self.user
        )

        importer.read_file()
        parsed_data = importer.parse_data()

        self.assertEqual(len(parsed_data['lines']), 1)
        self.assertIn('Café & Co.', parsed_data['lines'][0]['description'])







