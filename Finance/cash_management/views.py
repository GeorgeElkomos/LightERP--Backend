"""
Cash Management Views - API Endpoints
Provides REST API endpoints for PaymentType, Bank, BankBranch, BankAccount,
BankStatement, BankStatementLine, and BankStatementLineMatch operations.
"""
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import MultiPartParser, FormParser
from django.shortcuts import get_object_or_404
from django.core.exceptions import ValidationError
from django.db.models import Q, Sum
from django.http import HttpResponse
from decimal import Decimal
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import date

from .models import PaymentType, Bank, BankBranch, BankAccount, BankStatement, BankStatementLine, BankStatementLineMatch
from .serializers import (
    PaymentTypeListSerializer,
    PaymentTypeDetailSerializer,
    BankListSerializer,
    BankDetailSerializer,
    BankBranchListSerializer,
    BankBranchDetailSerializer,
    BankAccountListSerializer,
    BankAccountDetailSerializer,
    BankAccountBalanceUpdateSerializer,
    BankStatementListSerializer,
    BankStatementDetailSerializer,
    BankStatementCreateSerializer,
    BankStatementLineListSerializer,
    BankStatementLineDetailSerializer,
    BankStatementLineCreateSerializer,
    BankStatementLineMatchListSerializer,
    BankStatementLineMatchDetailSerializer,
    BankStatementLineMatchCreateSerializer,
    BankStatementImportSerializer,
    BankStatementImportPreviewSerializer,
)
from .services import BankStatementImporter


# ==================== PAYMENT TYPE VIEWSET ====================

class PaymentTypeViewSet(viewsets.ModelViewSet):
    """
    ViewSet for PaymentType CRUD operations.
    
    Endpoints:
    - GET /payment-types/ - List all payment types
    - POST /payment-types/ - Create new payment type
    - GET /payment-types/{id}/ - Get payment type details
    - PUT/PATCH /payment-types/{id}/ - Update payment type
    - DELETE /payment-types/{id}/ - Delete payment type
    """
    queryset = PaymentType.objects.all()
    permission_classes = [IsAuthenticated]
    
    def get_serializer_class(self):
        """Return appropriate serializer based on action"""
        if self.action == 'list':
            return PaymentTypeListSerializer
        return PaymentTypeDetailSerializer
    
    def get_queryset(self):
        """Filter queryset based on query parameters."""
        queryset = super().get_queryset()
        
        # Filter by active status
        is_active = self.request.query_params.get('is_active')
        if is_active is not None:
            is_active_bool = is_active.lower() == 'true'
            queryset = queryset.filter(is_active=is_active_bool)
        
        # Filter by enable_reconcile
        enable_reconcile = self.request.query_params.get('enable_reconcile')
        if enable_reconcile is not None:
            enable_reconcile_bool = enable_reconcile.lower() == 'true'
            queryset = queryset.filter(enable_reconcile=enable_reconcile_bool)
        
        # Search by name or code
        search = self.request.query_params.get('search')
        if search:
            queryset = queryset.filter(
                Q(payment_method_code__icontains=search) |
                Q(payment_method_name__icontains=search)
            )
        
        return queryset
    
    def perform_create(self, serializer):
        """Automatically set created_by to the current user"""
        serializer.save(created_by=self.request.user)


# ==================== BANK VIEWSET ====================

class BankViewSet(viewsets.ModelViewSet):
    """
    ViewSet for Bank CRUD operations.
    
    Endpoints:
    - GET /banks/ - List all banks
    - POST /banks/ - Create new bank
    - GET /banks/{id}/ - Get bank details
    - PUT/PATCH /banks/{id}/ - Update bank
    - DELETE /banks/{id}/ - Delete bank
    
    Custom Actions:
    - GET /banks/{id}/summary/ - Get bank statistics
    - POST /banks/{id}/activate/ - Activate bank
    - POST /banks/{id}/deactivate/ - Deactivate bank
    - GET /banks/{id}/branches/ - Get all branches
    - GET /banks/{id}/accounts/ - Get all accounts
    """
    queryset = Bank.objects.select_related('country').all()
    permission_classes = [IsAuthenticated]
    
    def get_serializer_class(self):
        """Return appropriate serializer based on action"""
        if self.action == 'list':
            return BankListSerializer
        return BankDetailSerializer
    
    def get_queryset(self):
        """
        Filter queryset based on query parameters.
        """
        queryset = super().get_queryset()
        
        # Filter by active status
        is_active = self.request.query_params.get('is_active')
        if is_active is not None:
            is_active_bool = is_active.lower() == 'true'
            queryset = queryset.filter(is_active=is_active_bool)
        
        # Filter by country
        country_id = self.request.query_params.get('country')
        if country_id:
            queryset = queryset.filter(country_id=country_id)
        
        country_code = self.request.query_params.get('country_code')
        if country_code:
            queryset = queryset.filter(country__code__iexact=country_code)
        
        # Search by name or code
        search = self.request.query_params.get('search')
        if search:
            queryset = queryset.filter(
                Q(bank_name__icontains=search) |
                Q(bank_code__icontains=search) |
                Q(swift_code__icontains=search)
            )
        
        return queryset
    
    def perform_create(self, serializer):
        """Create bank with user context"""
        serializer.save()
    
    def perform_update(self, serializer):
        """Update bank with user context"""
        serializer.save()
    
    @action(detail=True, methods=['get'])
    def summary(self, request, pk=None):
        """
        Get comprehensive bank statistics.
        
        GET /banks/{id}/summary/
        """
        bank = self.get_object()
        summary = bank.get_summary()
        return Response(summary, status=status.HTTP_200_OK)
    
    @action(detail=True, methods=['post'])
    def activate(self, request, pk=None):
        """
        Activate a bank.
        
        POST /banks/{id}/activate/
        """
        bank = self.get_object()
        bank.activate(user=request.user)
        return Response(
            {'message': f'Bank "{bank.bank_name}" activated successfully'},
            status=status.HTTP_200_OK
        )
    
    @action(detail=True, methods=['post'])
    def deactivate(self, request, pk=None):
        """
        Deactivate a bank.
        
        POST /banks/{id}/deactivate/
        """
        bank = self.get_object()
        result = bank.deactivate(user=request.user)
        return Response(result, status=status.HTTP_200_OK)
    
    @action(detail=True, methods=['get'])
    def branches(self, request, pk=None):
        """
        Get all branches for this bank.
        
        GET /banks/{id}/branches/
        Query params:
        - active_only: Return only active branches (true/false)
        """
        bank = self.get_object()
        active_only = request.query_params.get('active_only', 'false').lower() == 'true'
        
        branches = bank.get_branches(active_only=active_only)
        serializer = BankBranchListSerializer(branches, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
    
    @action(detail=True, methods=['get'])
    def accounts(self, request, pk=None):
        """
        Get all accounts across all branches for this bank.
        
        GET /banks/{id}/accounts/
        Query params:
        - active_only: Return only active accounts (true/false)
        """
        bank = self.get_object()
        active_only = request.query_params.get('active_only', 'false').lower() == 'true'
        
        accounts = bank.get_all_accounts(active_only=active_only)
        serializer = BankAccountListSerializer(accounts, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)


# ==================== BANK BRANCH VIEWSET ====================

class BankBranchViewSet(viewsets.ModelViewSet):
    """
    ViewSet for BankBranch CRUD operations.
    
    Endpoints:
    - GET /branches/ - List all branches
    - POST /branches/ - Create new branch
    - GET /branches/{id}/ - Get branch details
    - PUT/PATCH /branches/{id}/ - Update branch
    - DELETE /branches/{id}/ - Delete branch
    
    Custom Actions:
    - GET /branches/{id}/summary/ - Get branch statistics
    - POST /branches/{id}/activate/ - Activate branch
    - POST /branches/{id}/deactivate/ - Deactivate branch
    - GET /branches/{id}/accounts/ - Get all accounts
    - GET /branches/{id}/total_balance/ - Get total balance
    """
    queryset = BankBranch.objects.select_related('bank', 'country').all()
    permission_classes = [IsAuthenticated]
    
    def get_serializer_class(self):
        """Return appropriate serializer based on action"""
        if self.action == 'list':
            return BankBranchListSerializer
        return BankBranchDetailSerializer
    
    def get_queryset(self):
        """
        Filter queryset based on query parameters.
        """
        queryset = super().get_queryset()
        
        # Filter by bank
        bank_id = self.request.query_params.get('bank')
        if bank_id:
            queryset = queryset.filter(bank_id=bank_id)
        
        # Filter by active status
        is_active = self.request.query_params.get('is_active')
        if is_active is not None:
            is_active_bool = is_active.lower() == 'true'
            queryset = queryset.filter(is_active=is_active_bool)
        
        # Filter by country
        country_id = self.request.query_params.get('country')
        if country_id:
            queryset = queryset.filter(country_id=country_id)
        
        # Filter by city
        city = self.request.query_params.get('city')
        if city:
            queryset = queryset.filter(city__icontains=city)
        
        # Search
        search = self.request.query_params.get('search')
        if search:
            queryset = queryset.filter(
                Q(branch_name__icontains=search) |
                Q(branch_code__icontains=search) |
                Q(city__icontains=search) |
                Q(swift_code__icontains=search)
            )
        
        return queryset
    
    def perform_create(self, serializer):
        """Create branch with user context"""
        serializer.save()
    
    def perform_update(self, serializer):
        """Update branch with user context"""
        serializer.save()
    
    @action(detail=True, methods=['get'])
    def summary(self, request, pk=None):
        """
        Get comprehensive branch statistics.
        
        GET /branches/{id}/summary/
        """
        branch = self.get_object()
        summary = branch.get_summary()
        return Response(summary, status=status.HTTP_200_OK)
    
    @action(detail=True, methods=['post'])
    def activate(self, request, pk=None):
        """
        Activate a branch.
        
        POST /branches/{id}/activate/
        """
        branch = self.get_object()
        branch.activate(user=request.user)
        return Response(
            {'message': f'Branch "{branch.branch_name}" activated successfully'},
            status=status.HTTP_200_OK
        )
    
    @action(detail=True, methods=['post'])
    def deactivate(self, request, pk=None):
        """
        Deactivate a branch.
        
        POST /branches/{id}/deactivate/
        """
        branch = self.get_object()
        result = branch.deactivate(user=request.user)
        return Response(result, status=status.HTTP_200_OK)
    
    @action(detail=True, methods=['get'])
    def accounts(self, request, pk=None):
        """
        Get all accounts for this branch.
        
        GET /branches/{id}/accounts/
        Query params:
        - active_only: Return only active accounts (true/false)
        """
        branch = self.get_object()
        active_only = request.query_params.get('active_only', 'false').lower() == 'true'
        
        accounts = branch.get_accounts(active_only=active_only)
        serializer = BankAccountListSerializer(accounts, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
    
    @action(detail=True, methods=['get'])
    def total_balance(self, request, pk=None):
        """
        Get total balance across all accounts in this branch.
        
        GET /branches/{id}/total_balance/
        Query params:
        - currency: Filter by currency ID (optional)
        """
        branch = self.get_object()
        currency_id = request.query_params.get('currency')
        
        currency = None
        if currency_id:
            from Finance.core.models import Currency
            currency = get_object_or_404(Currency, pk=currency_id)
        
        total_balance = branch.get_total_balance(currency=currency)
        
        return Response({
            'branch': branch.branch_name,
            'branch_code': branch.branch_code,
            'total_balance': float(total_balance),
            'currency': currency.code if currency else 'All currencies',
        }, status=status.HTTP_200_OK)


# ==================== BANK ACCOUNT VIEWSET ====================

class BankAccountViewSet(viewsets.ModelViewSet):
    """
    ViewSet for BankAccount CRUD operations.
    
    Endpoints:
    - GET /accounts/ - List all accounts
    - POST /accounts/ - Create new account
    - GET /accounts/{id}/ - Get account details
    - PUT/PATCH /accounts/{id}/ - Update account
    - DELETE /accounts/{id}/ - Delete account
    
    Custom Actions:
    - POST /accounts/{id}/update_balance/ - Update account balance
    - POST /accounts/{id}/freeze/ - Freeze account
    - POST /accounts/{id}/unfreeze/ - Unfreeze account
    - GET /accounts/{id}/balance/ - Get balance summary
    - GET /accounts/{id}/check_balance/ - Check if sufficient balance
    - GET /accounts/{id}/hierarchy/ - Get full bank hierarchy
    """
    queryset = BankAccount.objects.select_related(
        'branch__bank',
        'branch__country',
        'currency',
        'cash_GL_combination',
        'cash_clearing_GL_combination'
    ).all()
    permission_classes = [IsAuthenticated]
    
    def get_serializer_class(self):
        """Return appropriate serializer based on action"""
        if self.action == 'list':
            return BankAccountListSerializer
        return BankAccountDetailSerializer
    
    def get_queryset(self):
        """
        Filter queryset based on query parameters.
        """
        queryset = super().get_queryset()
        
        # Filter by branch
        branch_id = self.request.query_params.get('branch')
        if branch_id:
            queryset = queryset.filter(branch_id=branch_id)
        
        # Filter by bank
        bank_id = self.request.query_params.get('bank')
        if bank_id:
            queryset = queryset.filter(branch__bank_id=bank_id)
        
        # Filter by active status
        is_active = self.request.query_params.get('is_active')
        if is_active is not None:
            is_active_bool = is_active.lower() == 'true'
            queryset = queryset.filter(is_active=is_active_bool)
        
        # Filter by currency
        currency_id = self.request.query_params.get('currency')
        if currency_id:
            queryset = queryset.filter(currency_id=currency_id)
        
        # Filter by account type
        account_type = self.request.query_params.get('account_type')
        if account_type:
            queryset = queryset.filter(account_type=account_type.upper())
        
        # Search
        search = self.request.query_params.get('search')
        if search:
            queryset = queryset.filter(
                Q(account_number__icontains=search) |
                Q(account_name__icontains=search) |
                Q(iban__icontains=search)
            )
        
        return queryset
    
    def perform_create(self, serializer):
        """Create account with user context"""
        serializer.save()
    
    def perform_update(self, serializer):
        """Update account with user context"""
        serializer.save()
    
    @action(detail=True, methods=['post'])
    def update_balance(self, request, pk=None):
        """
        Update account balance.
        
        POST /accounts/{id}/update_balance/
        Body:
        {
            "amount": 1000.00,
            "increase": true,
            "description": "Optional description"
        }
        """
        account = self.get_object()
        serializer = BankAccountBalanceUpdateSerializer(data=request.data)
        
        if serializer.is_valid():
            try:
                amount = serializer.validated_data['amount']
                increase = serializer.validated_data['increase']
                
                new_balance = account.update_balance(
                    amount=amount,
                    user=request.user,
                    increase=increase
                )
                
                return Response({
                    'message': 'Balance updated successfully',
                    'account_number': account.account_number,
                    'previous_balance': float(account.current_balance - (amount if increase else -amount)),
                    'amount': float(amount),
                    'increase': increase,
                    'new_balance': float(new_balance),
                }, status=status.HTTP_200_OK)
            
            except ValidationError as e:
                return Response(
                    {'error': str(e)},
                    status=status.HTTP_400_BAD_REQUEST
                )
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['post'])
    def freeze(self, request, pk=None):
        """
        Freeze (deactivate) an account.
        
        POST /accounts/{id}/freeze/
        """
        account = self.get_object()
        account.freeze(user=request.user)
        return Response(
            {'message': f'Account "{account.account_number}" frozen successfully'},
            status=status.HTTP_200_OK
        )
    
    @action(detail=True, methods=['post'])
    def unfreeze(self, request, pk=None):
        """
        Unfreeze (activate) an account.
        
        POST /accounts/{id}/unfreeze/
        """
        account = self.get_object()
        account.unfreeze(user=request.user)
        return Response(
            {'message': f'Account "{account.account_number}" unfrozen successfully'},
            status=status.HTTP_200_OK
        )
    
    @action(detail=True, methods=['get'])
    def balance(self, request, pk=None):
        """
        Get balance summary for account.
        
        GET /accounts/{id}/balance/
        """
        account = self.get_object()
        balance_summary = account.get_balance_summary()
        return Response(balance_summary, status=status.HTTP_200_OK)
    
    @action(detail=True, methods=['get'])
    def check_balance(self, request, pk=None):
        """
        Check if account has sufficient balance for a transaction.
        
        GET /accounts/{id}/check_balance/?amount=1000
        """
        account = self.get_object()
        amount_str = request.query_params.get('amount')
        
        if not amount_str:
            return Response(
                {'error': 'amount parameter is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            amount = Decimal(amount_str)
            sufficient = account.check_sufficient_balance(amount)
            
            return Response({
                'account_number': account.account_number,
                'current_balance': float(account.current_balance),
                'requested_amount': float(amount),
                'sufficient_balance': sufficient,
                'shortfall': float(amount - account.current_balance) if not sufficient else 0,
            }, status=status.HTTP_200_OK)
        
        except (ValueError, Decimal.InvalidOperation):
            return Response(
                {'error': 'Invalid amount format'},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=True, methods=['get'])
    def hierarchy(self, request, pk=None):
        """
        Get complete bank hierarchy (Bank → Branch → Account).
        
        GET /accounts/{id}/hierarchy/
        """
        account = self.get_object()
        hierarchy = account.get_full_hierarchy()
        return Response(hierarchy, status=status.HTTP_200_OK)


# ==================== BANK STATEMENT VIEWSET ====================

class BankStatementViewSet(viewsets.ModelViewSet):
    """
    ViewSet for BankStatement CRUD operations.
    
    Endpoints:
    - GET /statements/ - List all statements
    - POST /statements/ - Create new statement
    - GET /statements/{id}/ - Get statement details with lines
    - PUT/PATCH /statements/{id}/ - Update statement
    - DELETE /statements/{id}/ - Delete statement
    
    IMPORT Endpoints:
    - POST /statements/import_statement/ - Import bank statement from Excel/CSV
    - POST /statements/import_preview/ - Preview import without saving
    - GET /statements/download_template/ - Download Excel template
    """
    queryset = BankStatement.objects.select_related('bank_account', 'bank_account__branch', 'bank_account__branch__bank').all()
    permission_classes = [IsAuthenticated]
    
    def get_serializer_class(self):
        """Return appropriate serializer based on action"""
        if self.action == 'list':
            return BankStatementListSerializer
        elif self.action == 'create':
            return BankStatementCreateSerializer
        elif self.action == 'import_statement':
            return BankStatementImportSerializer
        elif self.action == 'import_preview':
            return BankStatementImportPreviewSerializer
        return BankStatementDetailSerializer
    
    def get_queryset(self):
        """Filter queryset based on query parameters."""
        queryset = super().get_queryset()
        
        # Filter by bank account
        bank_account_id = self.request.query_params.get('bank_account')
        if bank_account_id:
            queryset = queryset.filter(bank_account_id=bank_account_id)
        
        # Filter by reconciliation status
        reconciliation_status = self.request.query_params.get('reconciliation_status')
        if reconciliation_status:
            queryset = queryset.filter(reconciliation_status=reconciliation_status.upper())
        
        # Filter by date range
        date_from = self.request.query_params.get('date_from')
        if date_from:
            queryset = queryset.filter(statement_date__gte=date_from)
        
        date_to = self.request.query_params.get('date_to')
        if date_to:
            queryset = queryset.filter(statement_date__lte=date_to)
        
        return queryset
    
    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser, FormParser])
    def import_statement(self, request):
        """
        Import bank statement from Excel/CSV file.
        
        POST /statements/import_statement/
        
        Content-Type: multipart/form-data
        
        Fields:
        - file: Excel (.xlsx, .xls) or CSV (.csv) file
        - bank_account_id: ID of bank account
        - statement_number: Statement reference number
        - statement_date: Statement date (YYYY-MM-DD)
        - opening_balance: Opening balance (optional)
        - closing_balance: Closing balance (optional)
        
        HOW IT WORKS:
        1. Validates file format and size
        2. Parses Excel/CSV using pandas
        3. Maps columns flexibly (handles different bank formats)
        4. Validates all transaction data
        5. Creates BankStatement + all BankStatementLines
        6. Uses database transaction (all-or-nothing)
        7. Returns created statement with summary
        
        Returns:
        {
            "success": true,
            "message": "Statement imported successfully",
            "statement": { ... statement details ... },
            "lines_created": 25,
            "summary": {
                "total_rows": 25,
                "valid_rows": 25,
                "error_rows": 0,
                "total_debits": "5000.00",
                "total_credits": "7500.00"
            }
        }
        """
        serializer = self.get_serializer(data=request.data)
        
        if not serializer.is_valid():
            return Response(
                {
                    'success': False,
                    'errors': serializer.errors
                },
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # Initialize importer
            importer = BankStatementImporter(
                file_obj=serializer.validated_data['file'],
                bank_account_id=serializer.validated_data['bank_account_id'],
                user=request.user
            )
            
            # Import statement
            result = importer.import_statement({
                'statement_number': serializer.validated_data['statement_number'],
                'statement_date': serializer.validated_data['statement_date'],
                'opening_balance': serializer.validated_data.get('opening_balance', Decimal('0')),
                'closing_balance': serializer.validated_data.get('closing_balance', Decimal('0')),
            })
            
            # Serialize statement for response
            statement_serializer = BankStatementDetailSerializer(result['statement'])
            
            return Response({
                'success': True,
                'message': f'Statement imported successfully with {result["lines_created"]} transactions',
                'statement': statement_serializer.data,
                'lines_created': result['lines_created'],
                'summary': {
                    'total_rows': result['summary']['total_lines'],
                    'valid_rows': result['summary']['valid_lines'],
                    'error_rows': result['summary']['error_lines'],
                    'total_debits': str(result['summary']['total_debits']),
                    'total_credits': str(result['summary']['total_credits']),
                }
            }, status=status.HTTP_201_CREATED)
            
        except ValidationError as e:
            # Collect all errors from importer
            errors = importer.errors if hasattr(importer, 'errors') and importer.errors else [{'error': str(e)}]
            
            return Response({
                'success': False,
                'message': str(e),
                'errors': errors
            }, status=status.HTTP_400_BAD_REQUEST)
            
        except Exception as e:
            return Response({
                'success': False,
                'message': f'Import failed: {str(e)}',
                'errors': [{'error': str(e)}]
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser, FormParser])
    def import_preview(self, request):
        """
        Preview bank statement import without saving to database.
        
        POST /statements/import_preview/
        
        Content-Type: multipart/form-data
        
        Fields:
        - file: Excel (.xlsx, .xls) or CSV (.csv) file
        - bank_account_id: ID of bank account
        
        HOW IT WORKS:
        1. Parses file without creating any records
        2. Validates all data
        3. Returns preview of first 10 lines
        4. Shows any errors/warnings
        5. Shows totals and summary
        6. User can review before confirming actual import
        
        Returns:
        {
            "success": true,
            "statement_info": {
                "from_date": "2026-01-01",
                "to_date": "2026-01-31",
                "transaction_count": 25
            },
            "lines_preview": [ ... first 10 lines ... ],
            "total_lines": 25,
            "summary": {
                "total_rows": 25,
                "valid_rows": 25,
                "error_rows": 0,
                "total_debits": "5000.00",
                "total_credits": "7500.00"
            },
            "errors": [],
            "warnings": []
        }
        """
        serializer = self.get_serializer(data=request.data)
        
        if not serializer.is_valid():
            return Response(
                {
                    'success': False,
                    'errors': serializer.errors
                },
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # Initialize importer
            importer = BankStatementImporter(
                file_obj=serializer.validated_data['file'],
                bank_account_id=serializer.validated_data['bank_account_id'],
                user=request.user
            )
            
            # Preview import
            preview_data = importer.preview_import()
            
            return Response(preview_data, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response({
                'success': False,
                'message': f'Preview failed: {str(e)}',
                'errors': [{'error': str(e)}]
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['get'])
    def download_template(self, request):
        """
        Download Excel template for bank statement import.
        
        GET /statements/download_template/
        
        Returns an Excel file with:
        - Proper column headers
        - Sample data rows
        - Instructions sheet
        - Data validation
        
        HOW TO USE THE TEMPLATE:
        1. Download this template
        2. Fill in your bank statement data
        3. Keep column headers as-is
        4. Use YYYY-MM-DD format for dates
        5. Use numbers only for amounts (no currency symbols)
        6. Upload using import_statement endpoint
        """
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Create Instructions sheet
        ws_instructions = wb.active
        ws_instructions.title = "Instructions"
        
        # Add instructions
        instructions = [
            ["Bank Statement Import Template", ""],
            ["", ""],
            ["HOW TO USE:", ""],
            ["1. Go to 'Statement Data' sheet", ""],
            ["2. Fill in your transaction data", ""],
            ["3. Keep column headers as-is (row 1)", ""],
            ["4. Date format: YYYY-MM-DD (e.g., 2026-01-15)", ""],
            ["5. Amounts: Numbers only, no currency symbols", ""],
            ["6. Transaction Type: DEBIT or CREDIT", ""],
            ["7. Save file and upload via API", ""],
            ["", ""],
            ["REQUIRED COLUMNS:", ""],
            ["- Transaction Date", "Must be valid date"],
            ["- Description", "Cannot be empty"],
            ["- Amount OR Debit/Credit", "Must be numeric"],
            ["", ""],
            ["OPTIONAL COLUMNS:", ""],
            ["- Line Number", "Auto-generated if not provided"],
            ["- Value Date", "Defaults to Transaction Date"],
            ["- Reference Number", "Invoice/check number"],
            ["- Payee/Payer", "Who sent/received money"],
            ["- Balance", "Running balance after transaction"],
        ]
        
        for row_idx, row_data in enumerate(instructions, start=1):
            for col_idx, cell_value in enumerate(row_data, start=1):
                cell = ws_instructions.cell(row=row_idx, column=col_idx, value=cell_value)
                if row_idx == 1:
                    cell.font = Font(bold=True, size=14)
                elif row_data[0] in ["HOW TO USE:", "REQUIRED COLUMNS:", "OPTIONAL COLUMNS:"]:
                    cell.font = Font(bold=True)
        
        # Create Statement Data sheet
        ws_data = wb.create_sheet("Statement Data")
        
        # Define headers
        headers = [
            "Line Number",
            "Transaction Date",
            "Value Date",
            "Transaction Type",
            "Debit Amount",
            "Credit Amount",
            "Balance",
            "Reference Number",
            "Description",
            "Payee/Payer"
        ]
        
        # Add headers with styling
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_data.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Add sample data
        sample_data = [
            [1, "2026-01-15", "2026-01-15", "CREDIT", "", "5000.00", "15000.00", "DEP001", "Customer payment - Invoice #1001", "ABC Corporation"],
            [2, "2026-01-16", "2026-01-16", "DEBIT", "500.00", "", "14500.00", "CHQ001", "Rent payment", "Property Management Co"],
            [3, "2026-01-17", "2026-01-17", "CREDIT", "", "2500.00", "17000.00", "WIRE002", "Wire transfer received", "XYZ Ltd"],
        ]
        
        for row_idx, row_data in enumerate(sample_data, start=2):
            for col_idx, cell_value in enumerate(row_data, start=1):
                ws_data.cell(row=row_idx, column=col_idx, value=cell_value)
        
        # Auto-size columns
        for column in ws_data.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_data.column_dimensions[column_letter].width = adjusted_width
        
        # Save to response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=bank_statement_import_template_{date.today()}.xlsx'
        
        wb.save(response)
        return response


# ==================== BANK STATEMENT LINE VIEWSET ====================

class BankStatementLineViewSet(viewsets.ModelViewSet):
    """
    ViewSet for BankStatementLine CRUD operations.
    
    Endpoints:
    - GET /statement-lines/ - List all statement lines
    - POST /statement-lines/ - Create new statement line
    - GET /statement-lines/{id}/ - Get statement line details
    - PUT/PATCH /statement-lines/{id}/ - Update statement line
    - DELETE /statement-lines/{id}/ - Delete statement line
    """
    queryset = BankStatementLine.objects.select_related('bank_statement', 'matched_payment', 'reconciled_by').all()
    permission_classes = [IsAuthenticated]
    
    def get_serializer_class(self):
        """Return appropriate serializer based on action"""
        if self.action == 'list':
            return BankStatementLineListSerializer
        elif self.action == 'create':
            return BankStatementLineCreateSerializer
        return BankStatementLineDetailSerializer
    
    def get_queryset(self):
        """Filter queryset based on query parameters."""
        queryset = super().get_queryset()
        
        # Filter by bank statement
        bank_statement_id = self.request.query_params.get('bank_statement')
        if bank_statement_id:
            queryset = queryset.filter(bank_statement_id=bank_statement_id)
        
        # Filter by reconciliation status
        reconciliation_status = self.request.query_params.get('reconciliation_status')
        if reconciliation_status:
            queryset = queryset.filter(reconciliation_status=reconciliation_status.upper())
        
        # Filter by matched status
        has_match = self.request.query_params.get('has_match')
        if has_match is not None:
            if has_match.lower() == 'true':
                queryset = queryset.filter(matched_payment__isnull=False)
            elif has_match.lower() == 'false':
                queryset = queryset.filter(matched_payment__isnull=True)
        
        return queryset


# ==================== BANK STATEMENT LINE MATCH VIEWSET ====================

class BankStatementLineMatchViewSet(viewsets.ModelViewSet):
    """
    ViewSet for BankStatementLineMatch CRUD operations.
    
    Endpoints:
    - GET /matches/ - List all matches
    - POST /matches/ - Create new match
    - GET /matches/{id}/ - Get match details
    - PUT/PATCH /matches/{id}/ - Update match
    - DELETE /matches/{id}/ - Delete match (unmatch)
    """
    queryset = BankStatementLineMatch.objects.select_related('statement_line', 'payment', 'matched_by').all()
    permission_classes = [IsAuthenticated]
    
    def get_serializer_class(self):
        """Return appropriate serializer based on action"""
        if self.action == 'list':
            return BankStatementLineMatchListSerializer
        elif self.action == 'create':
            return BankStatementLineMatchCreateSerializer
        return BankStatementLineMatchDetailSerializer
    
    def get_queryset(self):
        """Filter queryset based on query parameters."""
        queryset = super().get_queryset()
        
        # Filter by statement line
        statement_line_id = self.request.query_params.get('statement_line')
        if statement_line_id:
            queryset = queryset.filter(statement_line_id=statement_line_id)
        
        # Filter by payment
        payment_id = self.request.query_params.get('payment')
        if payment_id:
            queryset = queryset.filter(payment_id=payment_id)
        
        # Filter by match status
        match_status = self.request.query_params.get('match_status')
        if match_status:
            queryset = queryset.filter(match_status=match_status.upper())
        
        # Filter by match type
        match_type = self.request.query_params.get('match_type')
        if match_type:
            queryset = queryset.filter(match_type=match_type.upper())
        
        return queryset
